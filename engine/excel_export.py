from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from data.classes import CLASSES
from data.teacher_loader import load_teachers
from data.labels import LABELS


def export_pretty_excel(class_table, filename="Pretty_Timetable.xlsx"):
    teachers = load_teachers()

    wb = Workbook()
    ws = wb.active
    ws.title = "Class Timetable"

    header = ["Class", "Day", "Period", "Subject", "Teacher"]
    ws.append(header)

    # Header style
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="ADD8E6")

    highlight_fill = PatternFill("solid", fgColor="90EE90")  # light green

    for cls, days in class_table.items():
        class_teacher_id = CLASSES[cls]["class_teacher"]
        class_teacher_name = teachers[class_teacher_id]["name"]

        for day, periods in days.items():
            for period, (subject, teacher_id) in periods.items():
                teacher_name = teachers[teacher_id]["name"]

                row = [
                    cls,
                    day,
                    period,
                    LABELS[subject]["en"],
                    teacher_name
                ]
                ws.append(row)

                # 🔍 Highlight P1 taught by class teacher
                if period == "P1" and teacher_id == class_teacher_id:
                    for cell in ws[ws.max_row]:
                        cell.fill = highlight_fill
                        cell.font = Font(bold=True)

    wb.save(filename)
